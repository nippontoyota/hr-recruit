from datetime import datetime, timezone
from io import BytesIO
from uuid import uuid4

from openpyxl import load_workbook
from fastapi.testclient import TestClient

from app.core.database import get_db
from app.core.deps import get_current_active_user
from app.main import app
from app.models.candidate import Candidate
from app.models.enums import PipelineStage, UserRole
from app.models.user import User
from app.services.candidate_export import build_candidates_workbook


client = TestClient(app)


def test_candidate_workbook_contains_roster_fields_and_is_safe_to_open():
    candidate = Candidate(
        id=uuid4(),
        candidate_id="NT-100",
        full_name="=Potential Formula",
        phone="9876543210",
        email="candidate@example.com",
        source="REFERRAL",
        position_applied_for="Sales Executive",
        experience="Fresher",
        current_stage=PipelineStage.SENT_TO_HO,
        branch_location="Kalamassery",
        applied_at=datetime(2026, 8, 31, 9, 42, tzinfo=timezone.utc),
    )

    workbook = load_workbook(build_candidates_workbook([candidate]))
    sheet = workbook["Candidates"]

    assert sheet.max_row == 2
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:S1"
    assert sheet["A1"].value == "Candidate ID"
    assert sheet["A2"].value == "NT-100"
    assert sheet["B2"].value == "'=Potential Formula"
    assert sheet["L2"].value == "SENT_TO_HO"


def test_candidate_export_is_head_office_only():
    local_user = User(
        id=uuid4(),
        email="branch@example.com",
        hashed_password="x",
        full_name="Branch HR",
        role=UserRole.LOCAL_HR,
        is_active=True,
    )
    app.dependency_overrides[get_current_active_user] = lambda: local_user
    try:
        response = client.get("/api/v1/candidates/export.xlsx")
        assert response.status_code == 403
    finally:
        app.dependency_overrides.clear()


def test_candidate_export_returns_a_valid_xlsx_for_head_office():
    ho_user = User(
        id=uuid4(),
        email="ho@example.com",
        hashed_password="x",
        full_name="Head Office HR",
        role=UserRole.HO_HR,
        is_active=True,
    )
    candidate = Candidate(
        id=uuid4(),
        candidate_id="NT-101",
        full_name="Valid Candidate",
        phone="9876543211",
        current_stage=PipelineStage.SENT_TO_HO,
    )

    class Result:
        def all(self):
            return [candidate]

    class FakeDb:
        def scalars(self, _query):
            return Result()

    app.dependency_overrides[get_current_active_user] = lambda: ho_user
    app.dependency_overrides[get_db] = lambda: FakeDb()
    try:
        response = client.get("/api/v1/candidates/export.xlsx")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        workbook = load_workbook(BytesIO(response.content))
        assert workbook["Candidates"]["A2"].value == "NT-101"
    finally:
        app.dependency_overrides.clear()
