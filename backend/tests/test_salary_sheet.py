from io import BytesIO
from types import SimpleNamespace
from uuid import uuid4

from openpyxl import Workbook
from pytest import raises

from app.core.salary_sheet import match_record, name_key, parse_salary_bytes


def _xlsx(build) -> bytes:
    wb = Workbook()
    build(wb.active)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _master(rows: list[dict]):
    headers = [
        "Job Code ",
        "Name",
        "Proposed DOJ",
        "Department",
        "Designation",
        "Branch",
        "Total Salary",
        "Total Allowance",
        "Fixed Incentive",
        "Gross Salary",
    ]

    def build(ws):
        for i, h in enumerate(headers, 1):
            ws.cell(1, i, h)
        for r, row in enumerate(rows, 2):
            for i, h in enumerate(headers, 1):
                ws.cell(r, i, row.get(h.strip()) or row.get(h))
    return _xlsx(build)


def test_rejects_individual_file():
    def build(ws):
        ws["A9"] = "SALARY PROPOSAL - NIPPON"
        ws["A10"] = "Name"
        ws["B10"] = "K S Mohasin"
        ws["A41"] = "GROSS SALARY"
        ws["B41"] = 15000
    with raises(ValueError, match="MASTER"):
        parse_salary_bytes(_xlsx(build))


def test_parse_master_and_name_punctuation():
    data = _master(
        [
            {
                "Name": "K.S. Mohasin",
                "Branch": "Kalamassery",
                "Total Salary": 12500,
                "Total Allowance": 0,
                "Fixed Incentive": 2500,
                "Gross Salary": 15000.0,
            }
        ]
    )
    fmt, records = parse_salary_bytes(data)
    assert fmt == "master"
    assert records[0]["gross salary"] == 15000
    assert name_key("K.S. Mohasin") == name_key("K S Mohasin")


def test_match_unique_name_and_ambiguous_branch():
    a = SimpleNamespace(
        id=uuid4(),
        full_name="K S Mohasin",
        current_stage="CSS",
        branch_location="Kalamassery",
        department="HR",
    )
    b = SimpleNamespace(
        id=uuid4(),
        full_name="K.S. Mohasin",
        current_stage="CSS",
        branch_location="Kollam",
        department="HR",
    )
    rec = {"name": "K S Mohasin", "branch": "Kalamassery"}
    hit, reason, collisions = match_record(rec, [a, b], {a.id, b.id})
    assert reason is None
    assert hit is a
    assert collisions == []

    ambiguous, why, collisions = match_record({"name": "K S Mohasin"}, [a, b], {a.id, b.id})
    assert ambiguous is None
    assert "Multiple" in why
    assert {c.id for c in collisions} == {a.id, b.id}


def test_skip_not_past_interviews_and_unknown_name():
    early = SimpleNamespace(
        id=uuid4(),
        full_name="K S Mohasin",
        current_stage="INTERVIEWS",
        branch_location="Kalamassery",
        department="HR",
    )
    rec = {"name": "K S Mohasin"}
    hit, reason, _ = match_record(rec, [early], set())
    assert hit is None
    assert "not passed" in reason

    missing, why, _ = match_record({"name": "Nobody"}, [early], {early.id})
    assert missing is None
    assert "No candidate" in why


def _person(**kwargs):
    defaults = dict(
        id=uuid4(),
        candidate_id="NT-1",
        full_name="Rahul Kumar",
        current_stage="CSS",
        branch_location="Kalamassery",
        department="HR",
        position_applied_for="Executive",
        email="rahul@test.com",
        phone="9876543210",
        salary_data=None,
    )
    defaults.update(kwargs)
    return SimpleNamespace(**defaults)


def test_match_portal_id_beats_duplicate_name():
    a = _person(candidate_id="NT-1", branch_location="Kalamassery")
    b = _person(id=uuid4(), candidate_id="NT-2", branch_location="Kollam")
    hit, reason, _ = match_record(
        {"name": "Rahul Kumar", "candidate id": "NT-2"},
        [a, b],
        {a.id, b.id},
    )
    assert reason is None
    assert hit is b


def test_legacy_portal_id_still_treated_as_id():
    a = _person(candidate_id="NT-2")
    hit, reason, _ = match_record(
        {"name": "Rahul Kumar", "candidate id": "NT-2026-00002"},
        [a],
        {a.id},
    )
    assert hit is None
    assert "NT-2026-00002" in (reason or "")


def test_match_rejects_id_when_name_disagrees():
    a = _person()
    hit, reason, _ = match_record(
        {"name": "Someone Else", "candidate id": "NT-1"},
        [a],
        {a.id},
    )
    assert hit is None
    assert "Someone Else" in reason


def test_match_phone_and_designation_tiebreak():
    a = _person(position_applied_for="Executive")
    b = _person(
        id=uuid4(),
        candidate_id="NT-2",
        phone="9999999999",
        position_applied_for="Advisor",
        email="other@test.com",
    )
    by_phone, _, _ = match_record(
        {"name": "Rahul Kumar", "phone": "9999999999"},
        [a, b],
        {a.id, b.id},
    )
    assert by_phone is b

    by_role, _, _ = match_record(
        {"name": "Rahul Kumar", "branch": "Kalamassery", "designation": "Advisor"},
        [a, b],
        {a.id, b.id},
    )
    assert by_role is b


def test_package_rejects_bad_math_and_yearly_mixup():
    from app.core.salary_sheet import validate_package

    assert validate_package({"name": "A", "total salary": 12500, "gross salary": 15000, "fixed incentive": 2500}) is None
    assert "add up" in (validate_package({"name": "A", "total salary": 12500, "gross salary": 20000}) or "")
    assert "yearly" in (validate_package({"name": "A", "total salary": 12500, "gross salary": 150000}) or "")
    assert validate_package({"name": "A"}) is not None


def test_unique_name_wrong_branch_is_not_auto_applied():
    a = _person(branch_location="Kollam")
    hit, reason, collisions = match_record(
        {"name": "Rahul Kumar", "branch": "Kalamassery", "total salary": 12500, "gross salary": 12500},
        [a],
        {a.id},
    )
    assert hit is None
    assert "Branch" in reason
    assert collisions == [a]


def test_review_suggests_close_name_and_locks_sent_offer():
    from app.core.salary_sheet import review_salary_records

    close = _person(full_name="K S Mohasin", candidate_id="NT-9")
    sent = _person(
        id=uuid4(),
        full_name="Anu Nithin",
        candidate_id="NT-10",
        offer_status="SENT",
        branch_location="Kalamassery",
    )
    proposed, skipped = review_salary_records(
        [
            {"name": "K S Mohasin", "total salary": 10000, "gross salary": 10000, "branch": "Kalamassery"},
            {"name": "K.S Mohasin", "total salary": 10000, "gross salary": 10000, "branch": "Kalamassery"},
            {"name": "Anu Nithin", "total salary": 20000, "gross salary": 20000, "branch": "Kalamassery"},
        ],
        [close, sent],
        {close.id, sent.id},
    )
    assert any(row["full_name"] == "K S Mohasin" for row in proposed)
    sent_skip = next(row for row in skipped if "Anu" in row["reason"] or row["name"] == "Anu Nithin")
    assert "sent" in sent_skip["reason"].lower()
