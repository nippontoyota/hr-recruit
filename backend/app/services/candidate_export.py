"""Build the Head Office candidate workbook."""

from __future__ import annotations

from datetime import date, datetime
from io import BytesIO, StringIO
from typing import Iterable, Iterator
import csv
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.candidate import Candidate


EXPORT_COLUMNS = (
    ("Candidate ID", "candidate_id"),
    ("Full name", "full_name"),
    ("Phone", "phone"),
    ("Email", "email"),
    ("Source", "source"),
    ("Source reference", "source_reference"),
    ("Position", "position_applied_for"),
    ("Experience", "experience"),
    ("Department", "department"),
    ("Opening type", "opening_type"),
    ("Branch", "branch_location"),
    ("Current stage", "current_stage"),
    ("Offer response", "offer_status"),
    ("Pre-form status", "pre_form_status"),
    ("Applied at", "applied_at"),
    ("Date added", "created_at"),
    ("Last updated", "updated_at"),
    ("Duplicate flagged", "is_duplicate_flagged"),
    ("Head Office hire", "is_head_office_hire"),
)

CSV_COLUMNS = EXPORT_COLUMNS[:16] + (("Application form sent", "pre_form_sent_at"),) + EXPORT_COLUMNS[16:]

_IST = ZoneInfo("Asia/Kolkata")


def _csv_value(value):
    if hasattr(value, "value"):
        value = value.value
    if isinstance(value, datetime):
        return value.astimezone(_IST).strftime("%d-%m-%Y") if value.tzinfo else value.strftime("%d-%m-%Y")
    if isinstance(value, date):
        return value.strftime("%d-%m-%Y")
    return "" if value is None else str(value)


def iter_candidates_csv(candidates: Iterable[Candidate]) -> Iterator[str]:
    buffer = StringIO()
    writer = csv.writer(buffer)
    writer.writerow([heading for heading, _ in CSV_COLUMNS])
    yield buffer.getvalue()
    for candidate in candidates:
        buffer.seek(0)
        buffer.truncate(0)
        writer.writerow([_csv_value(getattr(candidate, attribute, None)) for _, attribute in CSV_COLUMNS])
        yield buffer.getvalue()


def _cell_value(value):
    if hasattr(value, "value"):
        value = value.value
    if isinstance(value, datetime):
        # Excel stores datetimes without timezone information.
        return value.replace(tzinfo=None)
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value[:1] in ("=", "+", "-", "@"):
        # Prevent user-entered candidate data from being interpreted as a formula.
        return f"'{value}"
    return value


def build_candidates_workbook(candidates: Iterable[Candidate]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Candidates"
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(EXPORT_COLUMNS))}1"

    header_fill = PatternFill("solid", fgColor="0F172A")
    for column_index, (heading, _) in enumerate(EXPORT_COLUMNS, start=1):
        cell = sheet.cell(row=1, column=column_index, value=heading)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")

    for row_index, candidate in enumerate(candidates, start=2):
        for column_index, (_, attribute) in enumerate(EXPORT_COLUMNS, start=1):
            cell = sheet.cell(
                row=row_index,
                column=column_index,
                value=_cell_value(getattr(candidate, attribute, None)),
            )
            if isinstance(cell.value, (datetime, date)):
                cell.number_format = "dd-mm-yyyy hh:mm"

    widths = {
        "A": 16,
        "B": 28,
        "C": 16,
        "D": 30,
        "E": 16,
        "F": 24,
        "G": 24,
        "H": 14,
        "I": 18,
        "J": 16,
        "K": 18,
        "L": 24,
        "M": 16,
        "N": 16,
        "O": 21,
        "P": 21,
        "Q": 21,
        "R": 18,
        "S": 18,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
