"""Parse Nippon Salary Setting Sheet MASTER (one row per candidate)."""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from typing import Any

from openpyxl import load_workbook

_SPACE = re.compile(r"\s+")
_NON_ALNUM = re.compile(r"[^a-z0-9]+")
_NT_ID = re.compile(r"^NT-\d+(?:-\d+)?$", re.I)
_ID_KEYS = ("candidate id", "candidateid", "nt id", "portal id")
_EMAIL_KEYS = ("email", "e mail")
_PHONE_KEYS = ("phone", "phone number", "mobile number", "contact")

PAST_INTERVIEW_STAGES = frozenset(
    {"CSS", "SALARY_DETAILS", "FINAL_APPROVAL", "HIRED"}
)
BLOCKED_STAGES = frozenset({"REJECTED", "ON_HOLD"})


def fold_key(value: Any) -> str:
    text = str(value or "").strip().lower().replace("_", " ").replace("+", " ").replace(".", " ").replace("-", " ")
    text = text.replace(".", " ")
    return _SPACE.sub(" ", text).strip()


def name_key(value: Any) -> str:
    text = _NON_ALNUM.sub(" ", str(value or "").lower())
    tokens = [t for t in text.split() if t]
    return " ".join(sorted(tokens))


def json_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, bool):
        return value
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return int(value) if value.is_integer() else value
    if isinstance(value, int):
        return value
    text = str(value).strip()
    if not text or text.lower() in {"none", "nan", "n/a", "-"}:
        return None
    return text


def parse_salary_bytes(data: bytes) -> tuple[str, list[dict[str, Any]]]:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=False)
    records: list[dict[str, Any]] = []
    for sheet in wb.worksheets:
        records.extend(_parse_master(sheet))
    if not records:
        raise ValueError(
            "Not a MASTER salary sheet. First row must include Name and Gross Salary (or Total Salary)."
        )
    return "master", records


def _parse_master(sheet) -> list[dict[str, Any]]:
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [fold_key(h) or f"col_{i}" for i, h in enumerate(rows[0])]
    if "name" not in headers:
        return []
    if "gross salary" not in headers and "total salary" not in headers:
        return []
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if not raw or not any(raw):
            continue
        rec: dict[str, Any] = {}
        for key, cell in zip(headers, raw):
            value = json_cell(cell)
            if value is not None:
                rec[key] = value
        if rec.get("name"):
            out.append(rec)
    return out


def money(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        if value != value:
            return None
        return float(value)
    text = (
        str(value)
        .replace("Rs.", "")
        .replace("Rs", "")
        .replace("₹", "")
        .replace("/-", "")
        .replace(",", "")
        .strip()
    )
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def package_fields(record: dict[str, Any]) -> tuple[float | None, float, float, float | None]:
    total = money(record.get("total salary"))
    allowance = money(record.get("total allowance"))
    if allowance is None:
        allowance = 0.0
    others = money(record.get("fixed incentive"))
    if others is None:
        others = money(record.get("total incentive"))
    if others is None:
        others = money(record.get("others"))
    if others is None:
        others = 0.0
    gross = money(record.get("gross salary"))
    return total, allowance, others, gross


def validate_package(record: dict[str, Any]) -> str | None:
    if not record.get("name"):
        return "Row has no candidate name"
    total, allowance, others, gross = package_fields(record)
    if gross is None and total is None:
        return "Missing Total Salary and Gross Salary (formulas not saved? open in Excel and Save)"
    if total is not None and total <= 0:
        return "Total salary is zero"
    if gross is not None and gross <= 0:
        return "Gross salary is zero"
    if total is not None and gross is not None:
        if total > 0 and gross >= total * 9:
            return f"Gross {int(gross)} looks yearly vs monthly total {int(total)}"
        parts = total + allowance + others
        if abs(parts - gross) > max(1.0, gross * 0.02):
            return (
                f"Numbers don't add up: Total {int(total)} + Allowance {int(allowance)} "
                f"+ Incentive {int(others)} = {int(parts)}, but Gross is {int(gross)}"
            )
    return None


def normalize_package(record: dict[str, Any]) -> dict[str, Any]:
    out = dict(record)
    total, allowance, others, gross = package_fields(record)
    if gross is None and total is not None:
        out["gross salary"] = int(total + allowance + others) if (total + allowance + others).is_integer() else (total + allowance + others)
    if "others" not in out and others:
        out["others"] = int(others) if float(others).is_integer() else others
    return out


def suggest_names(sheet_name: Any, candidates: list[Any], limit: int = 5) -> list[Any]:
    from difflib import SequenceMatcher

    key = name_key(sheet_name)
    tokens = set(key.split())
    if not tokens:
        return []
    scored: list[tuple[float, Any]] = []
    for cand in candidates:
        other = name_key(getattr(cand, "full_name", ""))
        if not other:
            continue
        other_tokens = set(other.split())
        if tokens <= other_tokens or other_tokens <= tokens:
            ratio = 0.92
        else:
            ratio = SequenceMatcher(None, key, other).ratio()
        if ratio >= 0.78:
            scored.append((ratio, cand))
    scored.sort(key=lambda item: item[0], reverse=True)
    return [cand for _, cand in scored[:limit]]


def candidate_ref(cand: Any) -> dict[str, Any]:
    return {
        "id": str(cand.id),
        "candidate_id": getattr(cand, "candidate_id", "") or "",
        "full_name": cand.full_name,
        "branch": getattr(cand, "branch_location", None),
        "department": getattr(cand, "department", None),
    }


def is_past_interviews(stage: Any, selected: bool) -> bool:
    value = getattr(stage, "value", stage)
    value = str(value or "")
    if value in BLOCKED_STAGES:
        return False
    return selected or value in PAST_INTERVIEW_STAGES


def match_record(
    record: dict[str, Any],
    candidates: list[Any],
    selected_ids: set[Any],
    *,
    require_branch: bool = True,
) -> tuple[Any | None, str | None, list[Any]]:
    by_id = _match_unique_id(record, candidates)
    if by_id is not None:
        hit, err = by_id
        if err or hit is None:
            return None, err, []
        if not is_past_interviews(hit.current_stage, hit.id in selected_ids):
            return None, f"{hit.full_name} has not passed all interviews", []
        return hit, None, []

    sheet_name = record.get("name")
    key = name_key(sheet_name)
    if not key:
        return None, "Row has no candidate name", []

    named = [c for c in candidates if name_key(getattr(c, "full_name", "")) == key]
    if not named:
        return None, f"No candidate named {sheet_name}", []

    eligible = [
        c
        for c in named
        if is_past_interviews(c.current_stage, c.id in selected_ids)
    ]
    if not eligible:
        return None, f"{sheet_name} has not passed all interviews", named

    picked = _disambiguate(eligible, record)
    if picked is None:
        if len(eligible) == 1:
            only = eligible[0]
            return None, _branch_conflict(only, record) or f"Could not confirm {sheet_name}", eligible
        labels = ", ".join(
            f"{getattr(c, 'candidate_id', '')} ({getattr(c, 'branch_location', None) or 'no branch'})"
            for c in eligible
        )
        return (
            None,
            f"Multiple people named {sheet_name} ({labels}). Open the right profile and upload there.",
            eligible,
        )
    conflict = _branch_conflict(picked, record) if require_branch else None
    if conflict:
        return None, conflict, [picked]
    return picked, None, []


def review_salary_records(
    records: list[dict[str, Any]],
    candidates: list[Any],
    selected_ids: set[Any],
    pin: Any | None = None,
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    """Match + validate. Does not write. proposed items include the record to apply."""
    pool = [pin] if pin is not None else candidates
    proposed: list[dict[str, Any]] = []
    skipped: list[dict[str, Any]] = []
    seen: set[Any] = set()
    seen_jobs: set[str] = set()

    for record in records:
        bad = validate_package(record)
        if bad:
            skipped.append(_skip(record, bad, suggest_names(record.get("name"), candidates)))
            continue
        record = normalize_package(record)

        job = fold_key(record.get("job code"))
        if job:
            if job in seen_jobs:
                skipped.append(_skip(record, f"Duplicate Job Code {record.get('job code')} in this file"))
                continue
            seen_jobs.add(job)

        hit, reason, collisions = match_record(
            record, pool, selected_ids, require_branch=pin is None
        )
        if hit is None:
            hints = collisions or suggest_names(record.get("name"), candidates)
            skipped.append(_skip(record, reason or "Skipped", hints))
            continue
        if pin is not None and hit.id != pin.id:
            skipped.append(_skip(record, f"Sheet matched {hit.full_name}, not {pin.full_name}", [hit]))
            continue
        if str(getattr(hit, "offer_status", "") or "").upper() == "SENT":
            skipped.append(_skip(record, f"Offer already sent for {hit.full_name} — salary is locked", [hit]))
            continue
        if hit.id in seen:
            skipped.append(_skip(record, f"Duplicate row for {hit.full_name} — first match kept"))
            continue
        seen.add(hit.id)

        total, allowance, others, gross = package_fields(record)
        warnings: list[str] = []
        if getattr(hit, "salary_data", None):
            warnings.append("Replaces the salary already on file")
        sheet_branch = name_key(record.get("branch"))
        have_branch = name_key(getattr(hit, "branch_location", ""))
        if sheet_branch and not have_branch:
            warnings.append(f"Sheet branch is {record.get('branch')}; candidate has no branch on file")
        elif sheet_branch and have_branch and sheet_branch != have_branch:
            warnings.append(
                f"Branch on sheet is {record.get('branch')}, candidate is {hit.branch_location} (transfer?)"
            )

        proposed.append(
            {
                "id": str(hit.id),
                "candidate_id": hit.candidate_id,
                "full_name": hit.full_name,
                "branch": getattr(hit, "branch_location", None),
                "department": getattr(hit, "department", None),
                "total_salary": total,
                "total_allowance": allowance,
                "others": others,
                "gross_salary": gross,
                "joining_date": record.get("proposed doj"),
                "warnings": warnings,
                "record": record,
                "_candidate": hit,
            }
        )
    return proposed, skipped


def _skip(record: dict[str, Any], reason: str, matches: list[Any] | None = None) -> dict[str, Any]:
    return {
        "name": str(record.get("name") or ""),
        "reason": reason,
        "matches": [candidate_ref(c) for c in (matches or [])],
    }


def _branch_conflict(cand: Any, record: dict[str, Any]) -> str | None:
    want = name_key(record.get("branch"))
    have = name_key(getattr(cand, "branch_location", ""))
    if want and have and want != have:
        return (
            f"Branch on sheet is {record.get('branch')}, but "
            f"{getattr(cand, 'candidate_id', '')} is at {getattr(cand, 'branch_location', '')}. "
            "Open their profile and upload there if this is a transfer."
        )
    return None


def _match_unique_id(record: dict[str, Any], candidates: list[Any]) -> tuple[Any | None, str | None] | None:
    """Return (hit, err) if the sheet carries a unique id; None if it doesn't."""
    portal_id = _record_nt_id(record)
    if portal_id:
        hits = [c for c in candidates if str(getattr(c, "candidate_id", "")).upper() == portal_id]
        return _confirm_id_hit(hits, record, f"Candidate ID {portal_id}")

    email = _record_email(record)
    if email:
        hits = [c for c in candidates if (getattr(c, "email", None) or "").strip().lower() == email]
        return _confirm_id_hit(hits, record, email)

    phone = _record_phone(record)
    if phone:
        hits = [c for c in candidates if _phone_digits(getattr(c, "phone", None)) == phone]
        return _confirm_id_hit(hits, record, phone)

    job = fold_key(record.get("job code"))
    if job:
        hits = [
            c
            for c in candidates
            if fold_key((getattr(c, "salary_data", None) or {}).get("job code")) == job
        ]
        if hits:
            return _confirm_id_hit(hits, record, f"Job Code {record.get('job code')}")
    return None


def _confirm_id_hit(hits: list[Any], record: dict[str, Any], label: str) -> tuple[Any | None, str | None]:
    if not hits:
        return None, f"No candidate for {label}"
    if len(hits) > 1:
        return None, f"Multiple candidates for {label}"
    hit = hits[0]
    sheet_name = record.get("name")
    if sheet_name and name_key(sheet_name) != name_key(getattr(hit, "full_name", "")):
        return None, f"{label} is {hit.full_name}, but the sheet says {sheet_name}"
    return hit, None


def _record_nt_id(record: dict[str, Any]) -> str | None:
    for key in _ID_KEYS:
        text = str(record.get(key) or "").strip()
        if _NT_ID.match(text):
            return text.upper()
    job = str(record.get("job code") or "").strip()
    if _NT_ID.match(job):
        return job.upper()
    return None


def _record_email(record: dict[str, Any]) -> str | None:
    for key in _EMAIL_KEYS:
        text = str(record.get(key) or "").strip().lower()
        if "@" in text:
            return text
    return None


def _record_phone(record: dict[str, Any]) -> str | None:
    for key in _PHONE_KEYS:
        digits = _phone_digits(record.get(key))
        if digits:
            return digits
    return None


def _phone_digits(value: Any) -> str | None:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[-10:]
    return digits if len(digits) == 10 else None


def _disambiguate(candidates: list[Any], record: dict[str, Any]) -> Any | None:
    if len(candidates) == 1:
        return candidates[0]
    for sheet_key, cand_attr in (
        ("branch", "branch_location"),
        ("department", "department"),
        ("designation", "position_applied_for"),
    ):
        want = name_key(record.get(sheet_key))
        if not want:
            continue
        narrowed = [c for c in candidates if name_key(getattr(c, cand_attr, "")) == want]
        if len(narrowed) == 1:
            return narrowed[0]
        if narrowed:
            candidates = narrowed
    return None
